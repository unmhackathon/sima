import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Any

from config import EXCEL_FILE_PATH

logger = logging.getLogger(__name__)

REQUIRED_COLUMNS = {
    "number",
    "short_description",
    "description",
    "priority",
    "state",
    "sys_updated_on",
    "sys_created_on",
    "comments",
}

COLUMN_ALIASES = {
    "ticket_number": "number",
    "ticketid": "number",
    "title": "short_description",
    "summary": "short_description",
    "body": "description",
    "detail": "description",
    "comment": "comments",
    "notes": "comments",
    "updated_on": "sys_updated_on",
    "created_on": "sys_created_on",
    "created": "sys_created_on",
    "updated": "sys_updated_on",
    "status": "state",
    "activity_log": "activity_logs",
    "activitylogs": "activity_logs",
}


def parse_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value).strip()


def normalize_columns(columns: list[str]) -> dict[str, int]:
    normalized = {}
    for idx, name in enumerate(columns):
        if not name:
            continue
        key = str(name).strip().lower()
        normalized[key] = idx
        if key in COLUMN_ALIASES:
            normalized[COLUMN_ALIASES[key]] = idx
    return normalized


def load_tickets_from_excel(path: str) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        logger.error("openpyxl is required to read Excel files: %s", exc)
        raise

    workbook = None
    rows = []
    try:
        with open(path, "rb") as file_stream:
            workbook = load_workbook(file_stream, read_only=True, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
    finally:
        if workbook is not None:
            workbook.close()

    if not rows:
        return []

    headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
    column_index = {name: idx for idx, name in enumerate(headers) if name}
    for alias, target in COLUMN_ALIASES.items():
        if alias in column_index and target not in column_index:
            column_index[target] = column_index[alias]

    if not REQUIRED_COLUMNS.issubset(column_index):
        missing = REQUIRED_COLUMNS - set(column_index)
        raise ValueError(f"Excel file is missing required columns: {', '.join(sorted(missing))}")

    tickets = []
    for row in rows[1:]:
        if row is None or not any(row):
            continue

        activity_logs_idx = column_index.get("activity_logs")
        activity_logs = (
            parse_value(row[activity_logs_idx])
            if activity_logs_idx is not None and activity_logs_idx < len(row)
            else ""
        )

        ticket = {
            "sys_id": parse_value(row[column_index["number"]]),
            "number": parse_value(row[column_index["number"]]),
            "short_description": parse_value(row[column_index["short_description"]]),
            "description": parse_value(row[column_index["description"]]),
            "priority": parse_value(row[column_index["priority"]]),
            "state": parse_value(row[column_index["state"]]),
            "sys_updated_on": parse_value(row[column_index["sys_updated_on"]]),
            "sys_created_on": parse_value(row[column_index["sys_created_on"]]),
            "comments": parse_value(row[column_index["comments"]]),
            "activity_logs": activity_logs,
        }

        normalized_state = ticket["state"].strip().lower()
        if normalized_state in {"closed", "resolved", "cancelled", "canceled"}:
            continue

        tickets.append(ticket)

    return tickets


def sample_data() -> list[dict]:
    from datetime import datetime, timedelta, timezone

    now = datetime.now(timezone.utc)
    return [
        {
            "sys_id": "1",
            "number": "INC0010001",
            "short_description": "VPN is down for multiple users",
            "description": "Users cannot connect to corporate VPN since early morning.",
            "priority": "1",
            "state": "In Progress",
            "sys_updated_on": (now - timedelta(hours=50)).strftime("%Y-%m-%d %H:%M:%S"),
            "sys_created_on": (now - timedelta(hours=52)).strftime("%Y-%m-%d %H:%M:%S"),
            "comments": "The issue is affecting remote workers and needs immediate attention.",
        },
        {
            "sys_id": "2",
            "number": "INC0010002",
            "short_description": "Email delivery delay",
            "description": "Some users report email delivery delays but messages eventually arrive.",
            "priority": "3",
            "state": "On Hold",
            "sys_updated_on": (now - timedelta(hours=30)).strftime("%Y-%m-%d %H:%M:%S"),
            "sys_created_on": (now - timedelta(hours=35)).strftime("%Y-%m-%d %H:%M:%S"),
            "comments": "No recent update from the mail operations team.",
        },
        {
            "sys_id": "3",
            "number": "INC0010003",
            "short_description": "Application login error",
            "description": "Users see an authentication error while trying to access the HR portal.",
            "priority": "2",
            "state": "Active",
            "sys_updated_on": (now - timedelta(hours=10)).strftime("%Y-%m-%d %H:%M:%S"),
            "sys_created_on": (now - timedelta(hours=12)).strftime("%Y-%m-%d %H:%M:%S"),
            "comments": "The user says they are blocked from completing timesheet entries.",
        },
        {
            "sys_id": "4",
            "number": "INC0010004",
            "short_description": "Printer queue not clearing",
            "description": "One printer queue remains stuck and jobs are not printing.",
            "priority": "4",
            "state": "Active",
            "sys_updated_on": (now - timedelta(hours=2)).strftime("%Y-%m-%d %H:%M:%S"),
            "sys_created_on": (now - timedelta(hours=3)).strftime("%Y-%m-%d %H:%M:%S"),
            "comments": "Local site team is investigating.",
        },
    ]


def fetch_active_incidents() -> list[dict]:
    excel_path = Path(EXCEL_FILE_PATH)

    if excel_path.exists():
        try:
            tickets = load_tickets_from_excel(str(excel_path))
            logger.info("Loaded %d tickets from Excel: %s", len(tickets), excel_path)
            return tickets
        except Exception as exc:
            logger.warning("Could not read Excel file %s (%s); falling back to sample data.", excel_path, exc)

    logger.info("Ticket source not available; using sample ticket data.")
    return sample_data()
